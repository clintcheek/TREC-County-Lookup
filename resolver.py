from __future__ import annotations

import argparse
import csv
import hashlib
import html
import json
import os
import re
import signal
import threading
import time
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import asdict, dataclass, fields
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rapidfuzz.fuzz import token_set_ratio

RESOLVER_VERSION = "v7.7"
USER_AGENT = "TexasBrokerCountyResolver/7.7 (private license-location research)"
ADDRESS_RE = re.compile(
    r"\b(\d{1,6}\s+[A-Za-z0-9.'#&\- ]{2,80}?\s(?:Street|St|Road|Rd|Avenue|Ave|Boulevard|Blvd|Drive|Dr|Lane|Ln|Court|Ct|Parkway|Pkwy|Highway|Hwy|Loop|Way|Trail|Trl|Circle|Cir|Plaza|Place|Pl|Terrace|Ter)\.?"
    r"(?:\s*(?:Suite|Ste|Unit|#)\s*[A-Za-z0-9\-]+)?\s*,?\s*[A-Za-z.'\- ]{2,45},?\s*(?:TX|Texas|[A-Z]{2})\s+\d{5}(?:-\d{4})?)\b",
    re.I,
)
ZIP_RE = re.compile(r"\b\d{5}(?:-\d{4})?\b")
PHONE_RE = re.compile(r"(?<!\d)(?:\+?1[ .\-]?)?(?:\(\d{3}\)|\d{3})[ .\-]\d{3}[ .\-]\d{4}(?!\d)")
LOOSE_ADDRESS_RE = re.compile(
    r"\b(\d{1,6}\s+[A-Za-z0-9.'#&/\- ]{2,100}?,\s*[A-Za-z.'\- ]{2,45},?\s*(?:TX|Texas)\s+\d{5}(?:-\d{4})?)\b",
    re.I,
)
PROPERTY_TERMS = re.compile(
    r"\b(for sale|for lease|sold|pending|bed(?:room)?s?|bath(?:room)?s?|sq\.?\s*ft|acre(?:s)?|mls|listing price|property details|home value|open house)\b",
    re.I,
)
PROPERTY_PATHS = ("/property/", "/homedetail/", "/listing/", "/listings/", "/realestateandhomes-detail/", "/homes/", "/home-details/", "/real-estate/")
LISTING_DOMAINS = {"zillow.com", "redfin.com", "realtor.com", "homes.com", "har.com", "trulia.com", "loopnet.com", "crexi.com", "land.com", "landwatch.com"}
OFFICE_PATH_TERMS = ("contact", "about", "office", "location", "locations", "company", "team", "brokerage")
CONTACT_TERMS = re.compile(r"\b(contact|office|location|headquarters|address|about us|our office)\b", re.I)
CORP_SUFFIXES = {"llc", "lp", "ltd", "inc", "corp", "corporation", "company", "co", "pllc", "llp", "realty", "real", "estate", "group"}

_thread = threading.local()
_cache_lock = threading.Lock()
_audit_lock = threading.Lock()
_provider_fatal = threading.Event()
_provider_fatal_message = ""


class SearchProviderFatalError(RuntimeError):
    """A non-retryable search-provider failure that must stop the entire run."""


def set_provider_fatal(message: str) -> None:
    global _provider_fatal_message
    with _cache_lock:
        if not _provider_fatal.is_set():
            _provider_fatal_message = clean(message)
            _provider_fatal.set()


def provider_fatal_message() -> str:
    return _provider_fatal_message or "Search provider is unavailable."


def now_utc() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def session() -> requests.Session:
    if not hasattr(_thread, "session"):
        s = requests.Session()
        s.headers.update({"User-Agent": USER_AGENT, "Accept-Language": "en-US,en;q=0.9"})
        _thread.session = s
    return _thread.session


def clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def canonical_license(value: Any) -> str:
    return clean(value).upper()


def normalized_name(value: str) -> str:
    words = re.findall(r"[a-z0-9]+", clean(value).lower())
    kept = [w for w in words if w not in CORP_SUFFIXES]
    return " ".join(kept or words)


def normalize_address(value: str) -> str:
    x = clean(value).lower()
    replacements = {
        " street": " st", " road": " rd", " avenue": " ave", " boulevard": " blvd",
        " drive": " dr", " lane": " ln", " court": " ct", " parkway": " pkwy",
        " highway": " hwy", " suite ": " ste ", " texas ": " tx ",
    }
    for a, b in replacements.items():
        x = x.replace(a, b)
    return re.sub(r"[^a-z0-9]", "", x)


@dataclass
class Candidate:
    license_number: str
    brokerage_name: str
    related_broker_name: str
    address: str
    url: str
    source_domain: str
    source_tier: int
    source_type: str
    page_title: str
    evidence_text: str
    identity_score: float
    address_score: float
    negative_score: float
    total_score: float
    reject_reason: str = ""
    matched_address: str = ""
    city: str = ""
    state: str = ""
    zip_code: str = ""
    county: str = ""


@dataclass
class Result:
    license_number: str
    brokerage_name: str
    related_broker_name: str = ""
    office_address: str = ""
    city: str = ""
    state: str = ""
    zip_code: str = ""
    county: str = ""
    status: str = "Unresolved"
    confidence: float = 0.0
    identity_score: float = 0.0
    consensus_sources: int = 0
    evidence_url: str = ""
    secondary_evidence_url: str = ""
    evidence_type: str = ""
    notes: str = ""
    updated_at: str = ""
    resolver_version: str = RESOLVER_VERSION
    last_checked: str = ""
    evidence_count: int = 0
    needs_recheck: str = "Yes"
    attempt_count: int = 1
    brokerage_website: str = ""
    office_phone: str = ""



def finalize_result(result: Result, previous: Result | None = None) -> Result:
    result.resolver_version = RESOLVER_VERSION
    result.last_checked = result.updated_at or now_utc()
    result.updated_at = result.last_checked
    result.evidence_count = max(result.evidence_count, result.consensus_sources)
    result.needs_recheck = "No" if result.status == "Verified" and result.confidence >= 0.90 else "Yes"
    result.attempt_count = (previous.attempt_count if previous else 0) + 1
    return result


def status_rank(status: str) -> int:
    return {"Verified": 6, "Very Likely": 5, "Likely": 4, "Needs Review": 3, "Unresolved": 2, "Error": 1}.get(status, 0)


def choose_result(previous: Result | None, current: Result, force_replace: bool = False) -> Result:
    current = finalize_result(current, previous)
    if previous is None or force_replace:
        return current
    # Never downgrade a previously stronger answer. A new result replaces the old one
    # only when status improves, confidence improves materially, or evidence increases.
    enrichment_added = (
        (not previous.brokerage_website and bool(current.brokerage_website))
        or (not previous.office_phone and bool(current.office_phone))
        or (not previous.evidence_url and bool(current.evidence_url))
    )
    better = (
        status_rank(current.status) > status_rank(previous.status)
        or (status_rank(current.status) == status_rank(previous.status) and current.confidence > previous.confidence + 0.01)
        or (current.county == previous.county and current.evidence_count > previous.evidence_count and current.confidence >= previous.confidence - 0.02)
        or (current.county == previous.county and enrichment_added and current.confidence >= previous.confidence - 0.05)
    )
    if better:
        return current
    previous.last_checked = current.last_checked
    previous.attempt_count = current.attempt_count
    previous.resolver_version = RESOLVER_VERSION
    previous.needs_recheck = "No" if previous.status == "Verified" and previous.confidence >= 0.90 else "Yes"
    previous.notes = clean(previous.notes + " | Rechecked by " + RESOLVER_VERSION + "; existing stronger result retained.")
    return previous


def append_history(path: Path, previous: Result | None, attempted: Result, selected: Result, mode: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "checked_at", "mode", "resolver_version", "license_number", "brokerage_name",
        "previous_status", "previous_county", "previous_confidence",
        "attempted_status", "attempted_county", "attempted_confidence",
        "selected_status", "selected_county", "selected_confidence", "selection_note",
        "evidence_url", "secondary_evidence_url", "notes"
    ]
    row = {
        "checked_at": attempted.last_checked or attempted.updated_at or now_utc(),
        "mode": mode, "resolver_version": RESOLVER_VERSION,
        "license_number": attempted.license_number, "brokerage_name": attempted.brokerage_name,
        "previous_status": previous.status if previous else "",
        "previous_county": previous.county if previous else "",
        "previous_confidence": previous.confidence if previous else "",
        "attempted_status": attempted.status, "attempted_county": attempted.county,
        "attempted_confidence": attempted.confidence,
        "selected_status": selected.status, "selected_county": selected.county,
        "selected_confidence": selected.confidence,
        "selection_note": "new result selected" if selected is attempted else "existing stronger result retained",
        "evidence_url": attempted.evidence_url, "secondary_evidence_url": attempted.secondary_evidence_url,
        "notes": attempted.notes,
    }
    with _audit_lock:
        exists = path.exists()
        with path.open("a", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=fieldnames)
            if not exists:
                w.writeheader()
            w.writerow(row)


def registrable_host(url: str) -> str:
    host = urlparse(url).netloc.lower().split(":")[0].removeprefix("www.")
    return host


def is_listing_domain(url: str) -> bool:
    host = registrable_host(url)
    return any(host == d or host.endswith("." + d) for d in LISTING_DOMAINS)


NON_OFFICIAL_HOSTS = {
    "google.com", "facebook.com", "instagram.com", "linkedin.com", "x.com", "twitter.com",
    "yelp.com", "mapquest.com", "bbb.org", "bizapedia.com", "chamberofcommerce.com",
    "trec.texas.gov", "texas.gov", "har.com", "realtor.com", "zillow.com", "redfin.com",
    "homes.com", "trulia.com", "loopnet.com", "crexi.com", "youtube.com", "youtu.be",
}

def is_probable_official_website(url: str) -> bool:
    host = registrable_host(url)
    if not host or not url.lower().startswith(("http://", "https://")):
        return False
    return not any(host == d or host.endswith("." + d) for d in NON_OFFICIAL_HOSTS)

def extract_phones(text: str) -> list[str]:
    found: list[str] = []
    for raw in PHONE_RE.findall(clean(text)):
        digits = re.sub(r"\D", "", raw)
        if len(digits) == 11 and digits.startswith("1"):
            digits = digits[1:]
        if len(digits) != 10 or digits.startswith(("000", "555")):
            continue
        formatted = f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
        if formatted not in found:
            found.append(formatted)
    return found


def is_office_page(url: str, title: str = "", snippet: str = "") -> bool:
    path = urlparse(url).path.lower()
    sample = f"{title} {snippet}".lower()
    return any(term in path for term in OFFICE_PATH_TERMS) or bool(CONTACT_TERMS.search(sample))


def source_profile(url: str) -> tuple[int, str]:
    host = registrable_host(url)
    path = urlparse(url).path.lower()
    if any(p in path for p in PROPERTY_PATHS):
        return 0, "Property listing"
    if host.endswith("trec.texas.gov") or host.endswith("texas.gov"):
        return 5, "Government"
    # HAR/Realtor office, broker, agent and TREC profile pages are entity profiles,
    # not property listings, even though the host also carries listings.
    if host.endswith("har.com") and any(x in path for x in ("/broker_", "/agent_", "/trec_", "/real_estate_brokers", "/office_")):
        return 4, "Brokerage/agent profile"
    if host.endswith("realtor.com") and any(x in path for x in ("/realestateagency/", "/realestateagents/")):
        return 4, "Brokerage/agent profile"
    if is_listing_domain(url):
        return 1, "Listing-site corroboration only"
    if any(x in host for x in ("bbb.org", "chamberofcommerce.com")):
        return 3, "Business directory"
    if any(x in host for x in ("bizapedia.com", "mapquest.com", "yelp.com", "facebook.com", "instagram.com", "linkedin.com")):
        return 1, "Low-authority directory"
    return 4, "Possible official website"


def extract_addresses(text: str) -> list[str]:
    text = html.unescape(re.sub(r"\s+", " ", text or " "))
    out: list[str] = []
    for pattern in (ADDRESS_RE, LOOSE_ADDRESS_RE):
        for m in pattern.findall(text):
            a = clean(m).strip(" ,.;")
            # Reject obvious fragments while allowing numbered streets such as "2810 S. 27th".
            if len(a) < 12 or not ZIP_RE.search(a) or not re.search(r"\b(?:TX|Texas)\b", a, re.I):
                continue
            if a not in out:
                out.append(a)
    return out


def identity_score(name: str, broker_name: str, license_number: str, title: str, text: str, url: str) -> float:
    hay = clean(f"{title} {text} {url}").lower()
    name_score = token_set_ratio(normalized_name(name), hay) / 100
    broker_score = token_set_ratio(normalized_name(broker_name), hay) / 100 if broker_name else 0
    license_bonus = 0.35 if license_number.lower() in hay else 0
    exact_bonus = 0.20 if normalized_name(name) and normalized_name(name) in re.sub(r"[^a-z0-9 ]", " ", hay) else 0
    broker_bonus = 0.12 if broker_name and broker_score >= 0.90 else 0
    return min(1.0, 0.52 * name_score + 0.16 * broker_score + license_bonus + exact_bonus + broker_bonus)


def property_penalty(url: str, title: str, text: str, address: str) -> tuple[float, str]:
    path = urlparse(url).path.lower()
    sample = f"{title} {text}"
    if any(p in path for p in PROPERTY_PATHS):
        return 1.0, "property-listing URL"
    hits = len(PROPERTY_TERMS.findall(sample))
    if hits >= 3:
        return 0.85, "property-listing language"
    if hits >= 1 and address.lower() in sample.lower():
        return 0.45, "possible advertised property"
    return 0.0, ""


def address_context_score(title: str, text: str, source_tier: int) -> float:
    sample = f"{title} {text}"
    score = 0.30 + source_tier * 0.08
    if CONTACT_TERMS.search(sample):
        score += 0.18
    if ZIP_RE.search(sample):
        score += 0.08
    return min(1.0, score)


def read_search_cache(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_search_cache(path: Path, cache: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(".tmp")
    tmp.write_text(json.dumps(cache, ensure_ascii=False), encoding="utf-8")
    tmp.replace(path)


def serper_search(query: str, api_key: str, delay: float, cache: dict[str, Any], cache_path: Path, max_results: int) -> list[dict[str, str]]:
    key = hashlib.sha256(query.encode()).hexdigest()
    with _cache_lock:
        if key in cache:
            return cache[key]
    if _provider_fatal.is_set():
        raise SearchProviderFatalError(provider_fatal_message())
    if not clean(api_key):
        message = "SERPER_API_KEY is missing or empty"
        set_provider_fatal(message)
        raise SearchProviderFatalError(message)

    last_error = ""
    for attempt in range(3):
        if _provider_fatal.is_set():
            raise SearchProviderFatalError(provider_fatal_message())
        time.sleep(delay * (attempt + 1))
        try:
            r = session().post(
                "https://google.serper.dev/search",
                headers={"X-API-KEY": api_key, "Content-Type": "application/json"},
                json={"q": query, "gl": "us", "hl": "en", "num": max_results},
                timeout=40,
            )
            detail = clean(r.text)[:400]
            if r.status_code >= 400:
                lower = detail.lower()
                fatal = (
                    r.status_code in {401, 402, 403}
                    or "not enough credits" in lower
                    or "insufficient credit" in lower
                    or "invalid api key" in lower
                )
                message = f"Serper HTTP {r.status_code}: {detail}"
                if fatal:
                    set_provider_fatal(message)
                    raise SearchProviderFatalError(message)
                raise RuntimeError(message)
            payload = r.json()
            rows = [{
                "title": clean(i.get("title")), "link": clean(i.get("link")), "snippet": clean(i.get("snippet"))
            } for i in payload.get("organic", [])]
            with _cache_lock:
                cache[key] = rows
            return rows
        except SearchProviderFatalError:
            raise
        except Exception as exc:
            last_error = f"{type(exc).__name__}: {exc}"
            if attempt < 2:
                continue
    raise RuntimeError(last_error or "Unknown Serper search failure")


def fetch_page(url: str, delay: float) -> tuple[str, str, list[str], list[str], list[str], list[str]]:
    try:
        time.sleep(delay)
        r = session().get(url, timeout=30, allow_redirects=True)
        if r.status_code >= 400 or "text/html" not in r.headers.get("content-type", ""):
            return "", "", [], [], [], []
        soup = BeautifulSoup(r.text[:2_000_000], "lxml")
        title = clean(soup.title.get_text(" ", strip=True) if soup.title else "")
        text = clean(soup.get_text(" ", strip=True))[:120000]
        addresses: list[str] = []
        office_links: list[str] = []
        external_links: list[str] = []
        phones: list[str] = []
        base_host = registrable_host(r.url)
        for a in soup.find_all("a", href=True):
            label = clean(a.get_text(" ", strip=True)).lower()
            href = clean(a.get("href"))
            absolute = urljoin(r.url, href)
            linked_host = registrable_host(absolute)
            if linked_host != base_host:
                if is_probable_official_website(absolute) and absolute not in external_links:
                    external_links.append(absolute)
                continue
            path = urlparse(absolute).path.lower()
            if any(term in label or term in path for term in OFFICE_PATH_TERMS):
                if absolute not in office_links:
                    office_links.append(absolute)
        for node in soup.find_all("script", attrs={"type": "application/ld+json"}):
            try:
                obj = json.loads(node.get_text(" ", strip=True))
                stack = obj if isinstance(obj, list) else [obj]
                while stack:
                    item = stack.pop()
                    if isinstance(item, list):
                        stack.extend(item)
                    elif isinstance(item, dict):
                        stack.extend(item.values())
                        addr = item.get("address")
                        if isinstance(addr, dict):
                            line = ", ".join(clean(addr.get(k)) for k in ("streetAddress", "addressLocality", "addressRegion", "postalCode") if clean(addr.get(k)))
                            addresses.extend(extract_addresses(line))
            except Exception:
                pass
        addresses.extend(extract_addresses(text))
        phones.extend(extract_phones(text))
        for tel in soup.select("a[href^=\"tel:\"]"):
            phones.extend(extract_phones(clean(tel.get("href", ""))[4:]))
        return title, text, list(dict.fromkeys(addresses))[:15], office_links[:8], list(dict.fromkeys(phones))[:5], external_links[:10]
    except Exception:
        return "", "", [], [], [], []


def census_geocode(address: str, delay: float) -> dict[str, str] | None:
    try:
        time.sleep(delay)
        r = session().get(
            "https://geocoding.geo.census.gov/geocoder/geographies/onelineaddress",
            params={"address": address, "benchmark": "Public_AR_Current", "vintage": "Current_Current", "format": "json"},
            timeout=40,
        )
        r.raise_for_status()
        matches = r.json().get("result", {}).get("addressMatches", [])
        if not matches:
            return None
        m = matches[0]
        comps = m.get("addressComponents", {})
        counties = m.get("geographies", {}).get("Counties", [])
        county = clean(counties[0].get("NAME")) if counties else ""
        county = re.sub(r"\s+County$", "", county, flags=re.I)
        return {
            "matched_address": clean(m.get("matchedAddress")) or address,
            "city": clean(comps.get("city")), "state": clean(comps.get("state")),
            "zip_code": clean(comps.get("zip")), "county": county,
        }
    except Exception:
        return None


def build_candidate(lic: str, name: str, broker: str, address: str, url: str, title: str, text: str, source_type: str) -> Candidate:
    tier, profile = source_profile(url)
    ident = identity_score(name, broker, lic, title, text, url)
    neg, reason = property_penalty(url, title, text, address)
    addr_score = address_context_score(title, text, tier)
    total = max(0.0, min(1.0, 0.58 * ident + 0.27 * addr_score + 0.15 * (tier / 5) - 0.75 * neg))
    return Candidate(lic, name, broker, address, url, urlparse(url).netloc.lower(), tier, f"{profile}; {source_type}", title, clean(text)[:600], round(ident, 3), round(addr_score, 3), round(neg, 3), round(total, 3), reason)


def audit_candidates(path: Path, candidates: list[Candidate]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    names = [f.name for f in fields(Candidate)]
    with _audit_lock:
        exists = path.exists()
        with path.open("a", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=names)
            if not exists:
                w.writeheader()
            for c in candidates:
                w.writerow(asdict(c))


def relationship_score(candidate: Candidate) -> float:
    """Measure repeated person + brokerage identity, independent of the displayed address."""
    hay = re.sub(r"[^a-z0-9 ]", " ", clean(f"{candidate.page_title} {candidate.evidence_text} {candidate.url}").lower())
    firm = normalized_name(candidate.brokerage_name)
    person = normalized_name(candidate.related_broker_name)
    firm_exact = bool(firm and firm in hay)
    person_exact = bool(person and person in hay)
    license_hit = canonical_license(candidate.license_number).lower() in hay
    score = 0.0
    if firm_exact:
        score += 0.42
    if person_exact:
        score += 0.42
    if firm_exact and person_exact:
        score += 0.12
    if license_hit:
        score += 0.22
    return min(1.0, score)


def county_vote_weight(candidate: Candidate) -> float:
    """Convert source quality, identity and person/firm co-occurrence into a county vote."""
    tier_weight = {0: 0.12, 1: 0.24, 2: 0.42, 3: 0.68, 4: 0.88, 5: 1.0}.get(candidate.source_tier, 0.2)
    relation = relationship_score(candidate)
    office_bonus = 0.12 if is_office_page(candidate.url, candidate.page_title, candidate.evidence_text) else 0.0
    listing = is_listing_domain(candidate.url) or candidate.negative_score >= 0.80
    # Listing addresses are not office addresses, but repeated person+firm appearances
    # are useful operating-county evidence. Strong relationships receive a larger vote.
    listing_factor = (0.68 if relation >= 0.78 else 0.48 if relation >= 0.50 else 0.28) if listing else 1.0
    property_factor = 0.72 if listing and relation >= 0.78 else 0.50 if listing and relation >= 0.50 else max(0.12, 1.0 - candidate.negative_score)
    base = 0.36 * candidate.identity_score + 0.18 * candidate.address_score + 0.22 * tier_weight + 0.24 * relation + office_bonus
    return max(0.0, base * listing_factor * property_factor)


def confidence_label(score: float, authoritative: bool, independent_domains: int, margin: float, occurrences: int, strong_relationships: int) -> str:
    # v7.3 is optimized for usable lead generation rather than near-perfect registry reconstruction.
    if score >= 0.86 and margin >= 0.12 and (authoritative or (independent_domains >= 3 and strong_relationships >= 2)):
        return "Verified"
    if score >= 0.76 and margin >= 0.08 and independent_domains >= 2:
        return "Very Likely"
    if score >= 0.60 and margin >= 0.05 and (independent_domains >= 2 or occurrences >= 4):
        return "Likely"
    if score >= 0.42 and margin >= 0.02:
        return "Needs Review"
    return "Unresolved"


def build_query_stages(lic: str, name: str, broker: str, broker_license: str = "") -> list[tuple[str, list[str]]]:
    """Cheap deterministic entity resolution first; listing inference last."""
    negative_properties = "-site:zillow.com -site:redfin.com -site:trulia.com -site:loopnet.com -site:crexi.com"
    stages: list[tuple[str, list[str]]] = [
        ("exact_company", [
            f'"{name}" address Texas {negative_properties}',
            f'"{name}" (office OR contact OR location) Texas {negative_properties}',
        ]),
        ("company_license", [
            f'"{name}" "{lic}" Texas',
            f'"{name}" "{broker_license}" Texas' if broker_license else f'"{name}" Texas broker license',
        ]),
    ]
    if broker:
        stages.append(("person_company", [
            f'"{broker}" "{name}" Texas office address',
            f'"{broker}" "{name}" (broker OR realtor OR contact) Texas',
        ]))
        stages.append(("person_company_listings", [
            f'"{broker}" "{name}" listing Texas',
        ]))
        license_terms = " ".join(x for x in (lic, broker_license) if x)
        stages.append(("person_license_listings", [
            f'"{broker}" "{name}" {license_terms} listing Texas',
        ]))
    else:
        stages.append(("company_listings", [f'"{name}" listings Texas']))
    return stages


def build_queries(lic: str, name: str, broker: str, broker_license: str = "") -> list[str]:
    return [q for _, queries in build_query_stages(lic, name, broker, broker_license) for q in queries]


def exact_company_in_candidate(candidate: Candidate) -> bool:
    firm = normalized_name(candidate.brokerage_name)
    hay = re.sub(r"[^a-z0-9 ]", " ", clean(f"{candidate.page_title} {candidate.evidence_text} {candidate.url}").lower())
    return bool(firm and firm in hay)


def is_confirmed_office_candidate(candidate: Candidate) -> bool:
    if not exact_company_in_candidate(candidate):
        return False
    if candidate.negative_score >= 0.45 or any(p in urlparse(candidate.url).path.lower() for p in PROPERTY_PATHS):
        return False
    profile_signal = candidate.source_tier >= 4 or is_office_page(candidate.url, candidate.page_title, candidate.evidence_text)
    return profile_signal and candidate.identity_score >= 0.48


def gemini_grounded_address(name: str, lic: str, broker: str, broker_license: str, cfg: dict[str, Any]) -> dict[str, Any] | None:
    """Optional one-shot Google-grounded resolver for hard company-address cases."""
    api_key = os.environ.get("GEMINI_API_KEY", "").strip()
    if not api_key or not bool(cfg.get("enable_gemini_fallback", False)):
        return None
    model = clean(cfg.get("gemini_model", "gemini-3.5-flash"))
    prompt = f"""Find the Texas brokerage office or registered business address for this exact entity.
Company: {name}
Company license: {lic}
Related broker: {broker}
Individual broker license: {broker_license}
Prefer an official company contact page, HAR/TREC broker profile, regulatory disclosure, or an exact company-name business profile. Do not use a property listing address as the office. Return no address when the entity match is uncertain."""
    schema = {"type":"object","properties":{
        "entity_match":{"type":"boolean"},"matched_company_name":{"type":"string"},
        "office_address":{"type":"string"},"city":{"type":"string"},"state":{"type":"string"},
        "zip_code":{"type":"string"},"address_type":{"type":"string"},
        "confidence":{"type":"number"},"source_urls":{"type":"array","items":{"type":"string"}}
    },"required":["entity_match","matched_company_name","office_address","city","state","zip_code","address_type","confidence","source_urls"]}
    try:
        r = session().post(
            "https://generativelanguage.googleapis.com/v1beta/interactions",
            headers={"x-goog-api-key": api_key, "Content-Type": "application/json"},
            json={"model": model, "input": prompt, "tools":[{"type":"google_search"},{"type":"url_context"}],
                  "response_format":{"type":"text","mime_type":"application/json","schema":schema}},
            timeout=90,
        )
        r.raise_for_status()
        payload = r.json()
        text = clean(payload.get("output_text"))
        if not text:
            for item in payload.get("outputs", []) or payload.get("output", []):
                if isinstance(item, dict) and item.get("text"):
                    text = item["text"]; break
        data = json.loads(text) if text else None
        if not isinstance(data, dict) or not data.get("entity_match") or float(data.get("confidence") or 0) < 0.72:
            return None
        if clean(data.get("address_type")).lower() in {"listing", "property", "property_listing"}:
            return None
        return data
    except Exception as exc:
        print(f"Gemini fallback skipped: {type(exc).__name__}: {exc}", flush=True)
        return None

def resolve_one(lic: str, name: str, broker: str, broker_license: str, api_key: str, cfg: dict[str, Any], cache: dict[str, Any], cache_path: Path, audit_path: Path) -> Result:
    delay = float(cfg["request_delay_seconds"])
    stages = build_query_stages(lic, name, broker, broker_license)
    queries = [q for _, qs in stages for q in qs]
    candidates: list[Candidate] = []
    search_errors: list[str] = []
    successful_searches = 0
    seen_pages: set[str] = set()
    page_budget = int(cfg["max_pages_per_record"])
    contact_budget = int(cfg.get("max_contact_pages_per_record", 4))
    discovered_websites: list[str] = []
    discovered_phones: list[str] = []

    def remember_contact(urls: list[str] | None = None, phones: list[str] | None = None) -> None:
        for website in urls or []:
            if is_probable_official_website(website) and website not in discovered_websites:
                discovered_websites.append(website)
        for phone in phones or []:
            if phone and phone not in discovered_phones:
                discovered_phones.append(phone)

    def process_query(q: str) -> None:
        nonlocal successful_searches, page_budget, contact_budget
        try:
            rows = serper_search(q, api_key, delay, cache, cache_path, int(cfg["max_search_results"]))
            successful_searches += 1
        except SearchProviderFatalError:
            raise
        except Exception as exc:
            search_errors.append(f"{type(exc).__name__}: {exc}")
            return
        for row in rows:
            url, title, snippet = row["link"], row["title"], row["snippet"]
            if not url:
                continue
            remember_contact([url], extract_phones(f"{title} {snippet}"))
            for address in extract_addresses(f"{title} {snippet}"):
                candidates.append(build_candidate(lic, name, broker, address, url, title, snippet, "search snippet"))
            tier, _ = source_profile(url)
            ident = identity_score(name, broker, lic, title, snippet, url)
            is_property = any(p in urlparse(url).path.lower() for p in PROPERTY_PATHS)
            should_fetch = (not is_property and (is_office_page(url, title, snippet) or tier >= 3) and ident >= 0.30)
            if url not in seen_pages and page_budget > 0 and should_fetch:
                seen_pages.add(url); page_budget -= 1
                ptitle, ptext, addresses, office_links, page_phones, external_links = fetch_page(url, delay)
                remember_contact(external_links + ([url] if is_probable_official_website(url) else []), page_phones)
                for address in addresses:
                    candidates.append(build_candidate(lic, name, broker, address, url, ptitle or title, ptext or snippet, "office/business page"))
                for office_url in office_links:
                    if contact_budget <= 0 or office_url in seen_pages:
                        break
                    seen_pages.add(office_url); contact_budget -= 1
                    ctitle, ctext, caddresses, _, contact_phones, contact_external = fetch_page(office_url, delay)
                    remember_contact(contact_external + ([office_url] if is_probable_official_website(office_url) else []), contact_phones)
                    for address in caddresses:
                        candidates.append(build_candidate(lic, name, broker, address, office_url, ctitle, ctext, "linked contact/office page"))

    # Run cheap company-first stages and stop as soon as an exact company office is verified.
    for stage_name, stage_queries in stages:
        for q in stage_queries:
            process_query(q)
        if stage_name in {"exact_company", "company_license", "person_company"}:
            office_pool = sorted((c for c in candidates if is_confirmed_office_candidate(c)), key=lambda c: c.total_score, reverse=True)
            for c in office_pool[: int(cfg.get("max_office_geocodes_per_stage", 8))]:
                if not c.county:
                    geo = census_geocode(c.address, delay)
                    if geo and geo.get("county") and geo.get("state", "TX").upper() == "TX":
                        c.matched_address, c.city, c.state, c.zip_code, c.county = geo["matched_address"], geo["city"], geo["state"], geo["zip_code"], geo["county"]
                if c.county:
                    audit_candidates(audit_path, candidates[:40])
                    source_urls = [x.url for x in office_pool if x.url != c.url and registrable_host(x.url) != registrable_host(c.url)]
                    secondary = source_urls[0] if source_urls else ""
                    corroboration = len({registrable_host(x.url) for x in office_pool if x.county == c.county or not x.county})
                    conf = min(0.99, 0.90 + 0.02 * min(4, corroboration))
                    return Result(lic, name, broker, c.matched_address, c.city, c.state or "TX", c.zip_code, c.county,
                                  "Verified", round(conf,3), c.identity_score, max(1, corroboration), c.url, secondary,
                                  f"Exact company office; {c.source_type}",
                                  f"Exact company-name office resolution at stage {stage_name}; property listings were excluded from office-address selection.",
                                  now_utc(), evidence_count=max(1, corroboration),
                                  brokerage_website=(c.url if is_probable_official_website(c.url) else (discovered_websites[0] if discovered_websites else "")),
                                  office_phone=(discovered_phones[0] if discovered_phones else ""))
        if stage_name == "person_company":
            ai = gemini_grounded_address(name, lic, broker, broker_license, cfg)
            if ai:
                raw = clean(f"{ai.get('office_address')}, {ai.get('city')}, {ai.get('state') or 'TX'} {ai.get('zip_code')}")
                geo = census_geocode(raw, delay)
                if geo and geo.get("county"):
                    urls = [clean(x) for x in ai.get("source_urls", []) if clean(x)]
                    return Result(lic, name, broker, geo["matched_address"], geo["city"], geo["state"], geo["zip_code"], geo["county"],
                                  "Verified", min(0.96, float(ai.get("confidence") or 0.8)), 0.9, len({registrable_host(u) for u in urls}) or 1,
                                  urls[0] if urls else "", urls[1] if len(urls)>1 else "", "Gemini Google-grounded exact company office",
                                  "Google-grounded AI fallback found an exact company office after deterministic company searches did not produce a usable address; listing addresses were prohibited.",
                                  now_utc(), evidence_count=max(1,len(urls)),
                                  brokerage_website=next((u for u in urls if is_probable_official_website(u)), discovered_websites[0] if discovered_websites else ""),
                                  office_phone=(discovered_phones[0] if discovered_phones else ""))

    if successful_searches == 0:
        detail = " | ".join(dict.fromkeys(search_errors))[:1200]
        return Result(
            lic, name, broker, status="Error",
            notes="Search provider failed for every query; geocoding was not reached. " + detail,
            updated_at=now_utc(),
        )

    dedup: dict[tuple[str, str], Candidate] = {}
    for c in candidates:
        key = (normalize_address(c.address), c.source_domain)
        if key not in dedup or c.total_score > dedup[key].total_score:
            dedup[key] = c
    candidates = sorted(dedup.values(), key=lambda c: c.total_score, reverse=True)

    minimum_identity = float(cfg["minimum_identity_score"])
    eligible = [
        c for c in candidates
        if c.identity_score >= max(0.24, minimum_identity - 0.16)
        and (c.negative_score < 1.0 or relationship_score(c) >= 0.50)
    ]
    geocode_limit = int(cfg.get("max_geocodes_per_record", 24))
    for c in eligible[:geocode_limit]:
        geo = census_geocode(c.address, delay)
        if geo and geo.get("county") and geo.get("state", "TX").upper() == "TX":
            c.matched_address, c.city, c.state, c.zip_code, c.county = geo["matched_address"], geo["city"], geo["state"], geo["zip_code"], geo["county"]

    audit_candidates(audit_path, candidates[:40])
    geocoded = [c for c in eligible if c.county]
    if not geocoded:
        if not candidates:
            note = f"Search completed ({successful_searches}/{len(queries)} queries), but no Texas address candidates were extracted."
        elif not eligible:
            note = f"Found {len(candidates)} address candidate(s), but none passed the identity/property gates."
        else:
            note = f"Found {len(candidates)} candidate(s); {len(eligible)} reached geocoding, but Census returned no Texas county match."
        if search_errors:
            note += " Some searches also failed: " + " | ".join(dict.fromkeys(search_errors))[:500]
        return Result(lic, name, broker, status="Unresolved", notes=note, updated_at=now_utc())

    votes: dict[str, float] = defaultdict(float)
    county_candidates: dict[str, list[Candidate]] = defaultdict(list)
    county_domains: dict[str, set[str]] = defaultdict(set)
    county_occurrences: dict[str, int] = defaultdict(int)
    county_strong_relationships: dict[str, int] = defaultdict(int)

    # Aggregate repeated appearances with diminishing returns. A single website cannot
    # dominate, but several distinct listings on the same site still add useful evidence.
    by_county_domain: dict[tuple[str, str], list[Candidate]] = defaultdict(list)
    for c in geocoded:
        by_county_domain[(c.county, c.source_domain)].append(c)

    diminishing = (1.0, 0.38, 0.24, 0.16, 0.10)
    for (county_name, domain), items in by_county_domain.items():
        ranked_items = sorted(items, key=county_vote_weight, reverse=True)
        domain_vote = 0.0
        for index, c in enumerate(ranked_items[: len(diminishing)]):
            weight = county_vote_weight(c) * diminishing[index]
            if weight <= 0:
                continue
            domain_vote += weight
            county_candidates[county_name].append(c)
            county_occurrences[county_name] += 1
            if relationship_score(c) >= 0.78:
                county_strong_relationships[county_name] += 1
        if domain_vote > 0:
            votes[county_name] += domain_vote
            county_domains[county_name].add(domain)

    ranked = sorted(votes.items(), key=lambda x: x[1], reverse=True)
    if not ranked:
        return Result(lic, name, broker, status="Unresolved", notes="Evidence was found, but none met the minimum county-vote quality threshold.", updated_at=now_utc())

    top_county, top_vote = ranked[0]
    second_vote = ranked[1][1] if len(ranked) > 1 else 0.0
    total_vote = sum(votes.values()) or 1.0
    share = top_vote / total_vote
    margin = (top_vote - second_vote) / total_vote
    group = sorted(county_candidates[top_county], key=county_vote_weight, reverse=True)
    best = group[0]
    domains = county_domains[top_county]
    occurrences = county_occurrences[top_county]
    strong_relationships = county_strong_relationships[top_county]
    authoritative = any(c.source_tier >= 4 and c.negative_score < 0.45 for c in group)
    evidence_strength = min(1.0, top_vote / 1.45)
    density_strength = min(1.0, occurrences / 6)
    relationship_strength = min(1.0, strong_relationships / 3)
    confidence = min(0.99, 0.43 * share + 0.22 * evidence_strength + 0.15 * min(1.0, len(domains) / 3) + 0.10 * density_strength + 0.10 * relationship_strength)
    status = confidence_label(confidence, authoritative, len(domains), margin, occurrences, strong_relationships)

    non_listing = [c for c in group if c.negative_score < 0.45 and not any(p in urlparse(c.url).path.lower() for p in PROPERTY_PATHS)]
    listing_only = not non_listing
    # Repeated person+firm evidence across independent sites may reach Very Likely,
    # but listing-only evidence still cannot claim a verified office address.
    if listing_only and status == "Verified":
        status = "Very Likely"
        confidence = min(confidence, 0.85)

    county = top_county if status != "Unresolved" else ""
    secondary = next((c.url for c in group[1:] if c.source_domain != best.source_domain), "")
    inference_type = "Operating-county inference" if listing_only or best.negative_score >= 0.80 else "County evidence vote"
    notes = (
        f"{inference_type} selected {top_county}: vote share={share:.2f}, margin={margin:.2f}, "
        f"{len(domains)} independent domain(s), {occurrences} supporting occurrence(s), "
        f"{strong_relationships} strong person+brokerage relationship match(es), "
        f"authoritative_source={'yes' if authoritative else 'no'}. "
        "Repeated appearances were counted with diminishing returns; listing addresses were used as operating-area evidence, not presumed office addresses."
    )
    office_address = "" if listing_only or best.negative_score >= 0.80 else best.matched_address
    city = "" if listing_only or best.negative_score >= 0.80 else best.city
    state = "TX" if county else ""
    zip_code = "" if listing_only or best.negative_score >= 0.80 else best.zip_code
    return Result(
        lic, name, broker, office_address, city, state, zip_code, county,
        status, round(confidence, 3), best.identity_score, len(domains), best.url, secondary,
        inference_type + "; " + best.source_type, notes, now_utc(), evidence_count=occurrences,
        brokerage_website=(best.url if is_probable_official_website(best.url) else (discovered_websites[0] if discovered_websites else "")),
        office_phone=(discovered_phones[0] if discovered_phones else ""),
    )

def find_columns(ws) -> tuple[int, int, int, int | None, int | None]:
    headers = {clean(c.value).lower(): c.column for c in ws[1]}
    lic = headers.get("license number") or headers.get("license")
    name = headers.get("full name") or headers.get("brokerage") or headers.get("brokerage name")
    broker = headers.get("related license full name") or headers.get("related broker name")
    broker_license = headers.get("related license number") or headers.get("individual broker number")
    if not lic or not name:
        raise RuntimeError(f"Required columns not found: {list(headers)}")
    return 1, lic, name, broker, broker_license


def load_checkpoint(path: Path) -> dict[str, Result]:
    if not path.exists():
        return {}
    out: dict[str, Result] = {}
    names = {f.name for f in fields(Result)}
    with path.open(newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            data = {k: row.get(k, "") for k in names}
            for k in ("confidence", "identity_score"):
                data[k] = float(data.get(k) or 0)
            data["consensus_sources"] = int(float(data.get("consensus_sources") or 0))
            data["evidence_count"] = int(float(data.get("evidence_count") or data.get("consensus_sources") or 0))
            data["attempt_count"] = int(float(data.get("attempt_count") or 0))
            r = Result(**data)
            out[canonical_license(r.license_number)] = r
    return out


def save_checkpoint(path: Path, results: dict[str, Result]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(".tmp")
    names = [f.name for f in fields(Result)]
    with tmp.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=names); w.writeheader()
        for key in sorted(results): w.writerow(asdict(results[key]))
    tmp.replace(path)


def write_output(input_file: Path, output_file: Path, checkpoint: dict[str, Result]) -> None:
    wb = load_workbook(input_file)
    ws = wb.active
    _, lic_col, _, _, _ = find_columns(ws)
    headers = ["Brokerage Website", "Office Phone", "Office Address", "City", "State", "ZIP", "County", "Resolution Status", "Confidence", "Identity Score", "Consensus Sources", "Evidence Type", "Evidence URL", "Secondary Evidence URL", "Resolution Notes", "Last Updated UTC", "Resolver Version", "Last Checked UTC", "Evidence Count", "Needs Recheck", "Attempt Count"]
    existing = {clean(c.value): c.column for c in ws[1]}
    col = ws.max_column + 1
    for h in headers:
        if h not in existing:
            ws.cell(1, col, h); existing[h] = col; col += 1
    mapping = {
        "Brokerage Website":"brokerage_website", "Office Phone":"office_phone",
        "Office Address":"office_address", "City":"city", "State":"state", "ZIP":"zip_code", "County":"county",
        "Resolution Status":"status", "Confidence":"confidence", "Identity Score":"identity_score",
        "Consensus Sources":"consensus_sources", "Evidence Type":"evidence_type", "Evidence URL":"evidence_url",
        "Secondary Evidence URL":"secondary_evidence_url", "Resolution Notes":"notes", "Last Updated UTC":"updated_at",
        "Resolver Version":"resolver_version", "Last Checked UTC":"last_checked", "Evidence Count":"evidence_count",
        "Needs Recheck":"needs_recheck", "Attempt Count":"attempt_count",
    }
    for row in range(2, ws.max_row + 1):
        r = checkpoint.get(canonical_license(ws.cell(row, lic_col).value))
        if r:
            for h, attr in mapping.items(): ws.cell(row, existing[h], getattr(r, attr))
    fill = PatternFill("solid", fgColor="1F4E78")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = fill; c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = {"Brokerage Website":42,"Office Phone":18,"Office Address":38,"City":18,"State":9,"ZIP":12,"County":18,"Resolution Status":18,"Confidence":12,"Identity Score":14,"Consensus Sources":18,"Evidence Type":26,"Evidence URL":52,"Secondary Evidence URL":52,"Resolution Notes":55,"Last Updated UTC":22,"Resolver Version":16,"Last Checked UTC":22,"Evidence Count":15,"Needs Recheck":15,"Attempt Count":14}
    for h, width in widths.items(): ws.column_dimensions[get_column_letter(existing[h])].width = width
    ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions

    # A concise, ready-to-use directory with only the fields requested for business use.
    if "Clean Results" in wb.sheetnames:
        del wb["Clean Results"]
    clean_ws = wb.create_sheet("Clean Results", 0)
    clean_headers = [
        "License Number", "Brokerage Name", "Related Broker Name", "Office Address",
        "City", "State", "ZIP", "County", "Office Phone", "Brokerage Website",
        "Verification Website 1", "Verification Website 2", "Resolution Status", "Confidence"
    ]
    clean_ws.append(clean_headers)
    for key in sorted(checkpoint):
        r = checkpoint[key]
        clean_ws.append([
            r.license_number, r.brokerage_name, r.related_broker_name, r.office_address,
            r.city, r.state, r.zip_code, r.county, r.office_phone, r.brokerage_website,
            r.evidence_url, r.secondary_evidence_url, r.status, r.confidence
        ])
    for c in clean_ws[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    clean_widths = [16,34,28,40,18,9,12,18,18,42,52,52,18,12]
    for idx, width in enumerate(clean_widths, 1):
        clean_ws.column_dimensions[get_column_letter(idx)].width = width
    clean_ws.freeze_panes = "A2"
    clean_ws.auto_filter.ref = clean_ws.dimensions

    output_file.parent.mkdir(parents=True, exist_ok=True)
    temp_output = output_file.with_name(output_file.stem + ".checkpoint" + output_file.suffix)
    wb.save(temp_output)
    temp_output.replace(output_file)


def version_tuple(value: str) -> tuple[int, ...]:
    numbers = tuple(int(x) for x in re.findall(r"\d+", clean(value)))
    return numbers or (0,)


def age_days(timestamp: str) -> int:
    try:
        dt = datetime.fromisoformat(timestamp.replace("Z", "+00:00"))
        return max(0, (datetime.now(timezone.utc) - dt).days)
    except Exception:
        return 999999


def parallel_queue_for(old: Result | None, related_broker: str) -> str:
    """Assign each record to exactly one of the five v7.6 queues.

    Priority prevents duplicate work when a record qualifies for several queues.
    """
    if old is None:
        return "new"
    state = clean(old.state).upper()
    county = clean(old.county)
    status = clean(old.status)
    needs_recheck = clean(old.needs_recheck).lower() in {"yes", "true", "1"}

    # A company with an associated/related broker but still no county is the
    # most actionable repair queue after genuinely new records.
    if clean(related_broker) and not county:
        return "broker_no_county"
    if state and state not in {"TX", "TEXAS"}:
        return "out_of_state"
    if status in {"Unresolved", "Error"}:
        return "unresolved"
    if status in {"Very Likely", "Likely", "Needs Review"} or needs_recheck:
        return "review"
    return ""


def should_process(old: Result | None, mode: str, cfg: dict[str, Any], related_broker: str = "") -> bool:
    if mode in {"parallel_15", "parallel_5"}:
        return bool(parallel_queue_for(old, related_broker))
    if old is None:
        return True
    threshold = float(cfg.get("low_confidence_threshold", 0.90))
    stale_days = int(cfg.get("recheck_after_days", 180))
    if mode == "recheck_all":
        return True
    if mode in {"recheck_review", "recheck_unresolved"}:
        return old.status in ({"Very Likely", "Likely", "Needs Review", "Unresolved"} if mode == "recheck_review" else {"Unresolved", "Error"})
    if mode == "upgrade_confidence":
        return old.confidence < threshold or old.status != "Verified"
    if mode == "upgrade_version":
        return version_tuple(old.resolver_version) < version_tuple(RESOLVER_VERSION)
    if mode == "recheck_stale":
        return age_days(old.last_checked or old.updated_at) >= stale_days
    if mode == "flagged":
        return clean(old.needs_recheck).lower() in {"yes", "true", "1"}
    if mode == "new" and old.status == "Error":
        note = old.notes.lower()
        return any(term in note for term in ("search provider", "serper", "api_key", "credits"))
    return False


def save_progress_metadata(path: Path, checkpoint: dict[str, Result], completed_this_run: int, planned_this_run: int, mode: str) -> None:
    counts: dict[str, int] = defaultdict(int)
    for result in checkpoint.values():
        counts[result.status] += 1
    payload = {
        "last_saved_utc": now_utc(),
        "completed_this_run": completed_this_run,
        "planned_this_run": planned_this_run,
        "total_checkpointed_records": len(checkpoint),
        "mode": mode,
        "resolver_version": RESOLVER_VERSION,
        "status_counts": dict(counts),
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    temp = path.with_suffix(".tmp")
    temp.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    temp.replace(path)


def persist_all(
    checkpoint_path: Path,
    metadata_path: Path,
    cache_path: Path,
    input_file: Path,
    output_file: Path,
    checkpoint: dict[str, Result],
    cache: dict[str, Any],
    completed_this_run: int,
    planned_this_run: int,
    mode: str,
) -> None:
    """Atomically persist CSV state, cache, metadata, and the current workbook."""
    save_checkpoint(checkpoint_path, checkpoint)
    save_search_cache(cache_path, cache)
    save_progress_metadata(metadata_path, checkpoint, completed_this_run, planned_this_run, mode)
    write_output(input_file, output_file, checkpoint)
    print(f"Saved recovery checkpoint: {completed_this_run}/{planned_this_run} completed this run; {len(checkpoint)} total records saved.", flush=True)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--config", default="config.json")
    ap.add_argument("--max-rows", type=int)
    ap.add_argument("--mode", choices=["parallel_15", "parallel_5", "new", "recheck_review", "recheck_unresolved", "upgrade_confidence", "upgrade_version", "recheck_stale", "flagged", "recheck_all"], default="parallel_15")
    ap.add_argument("--workers", type=int, default=None, help="Concurrent resolver workers; v7.7 defaults to 15.")
    ap.add_argument("--checkpoint-every", type=int, default=None)
    args = ap.parse_args()

    cfg = json.loads(Path(args.config).read_text(encoding="utf-8"))
    api_key = os.environ.get("SERPER_API_KEY", "").strip()
    if not api_key:
        raise SystemExit("SERPER_API_KEY is missing.")

    input_file = Path(cfg["input_file"])
    output_file = Path(cfg["output_file"])
    checkpoint_path = Path(cfg["checkpoint_file"])
    audit_path = Path(cfg["candidate_audit_file"])
    cache_path = Path(cfg["search_cache_file"])
    metadata_path = Path(cfg.get("checkpoint_metadata_file", "state/checkpoint.json"))
    history_path = Path(cfg.get("history_file", "state/resolution_history.csv"))
    checkpoint_every = max(1, args.checkpoint_every or int(cfg.get("checkpoint_every", 25)))

    checkpoint = load_checkpoint(checkpoint_path)
    cache = read_search_cache(cache_path)

    wb = load_workbook(input_file, read_only=True, data_only=True)
    ws = wb.active
    _, lic_col, name_col, broker_col, broker_license_col = find_columns(ws)
    pending: list[tuple[str, str, str, str]] = []
    queue_counts: dict[str, int] = defaultdict(int)
    limit = args.max_rows or int(cfg["max_rows_per_run"])
    for row in ws.iter_rows(min_row=2, values_only=True):
        lic = canonical_license(row[lic_col - 1])
        name = clean(row[name_col - 1])
        broker = clean(row[broker_col - 1]) if broker_col else ""
        broker_license = canonical_license(row[broker_license_col - 1]) if broker_license_col else ""
        old = checkpoint.get(lic)
        if lic and name and should_process(old, args.mode, cfg, broker):
            pending.append((lic, name, broker, broker_license))
            if args.mode in {"parallel_15", "parallel_5"}:
                queue_counts[parallel_queue_for(old, broker)] += 1
            if len(pending) >= limit:
                break
    wb.close()

    planned = len(pending)
    completed = 0
    interrupted = False
    run_started = time.monotonic()
    last_progress_print = run_started

    def handle_stop(signum, frame):
        nonlocal interrupted
        interrupted = True
        print(f"Received stop signal {signum}. Finishing current completed records and saving progress.", flush=True)

    signal.signal(signal.SIGTERM, handle_stop)
    signal.signal(signal.SIGINT, handle_stop)

    worker_count = min(32, max(1, args.workers or int(cfg.get("workers", 15))))
    print(
        f"Checkpoint contains {len(checkpoint)} records; mode={args.mode}; processing={planned}; "
        f"workers={worker_count}; save interval={checkpoint_every}.",
        flush=True,
    )
    if args.mode in {"parallel_15", "parallel_5"}:
        print(f"v7.7 priority plan: {dict(queue_counts)}", flush=True)

    # Create a valid recovery workbook and metadata even when there is nothing new to process.
    if not pending:
        persist_all(
            checkpoint_path, metadata_path, cache_path, input_file, output_file,
            checkpoint, cache, completed, planned, args.mode,
        )
        print("No eligible records remain for this mode.", flush=True)
        return 0

    # Preflight uses the first real query. A successful response is cached and reused,
    # so this does not spend an extra search credit. Fatal account/key failures stop
    # before any record is marked Error or counted as completed.
    first_lic, first_name, first_broker, first_broker_license = pending[0]
    try:
        serper_search(
            build_queries(first_lic, first_name, first_broker, first_broker_license)[0],
            api_key,
            float(cfg["request_delay_seconds"]),
            cache,
            cache_path,
            int(cfg["max_search_results"]),
        )
    except SearchProviderFatalError as exc:
        print(f"::error::Search-provider preflight failed: {exc}", flush=True)
        print("No broker records were changed. Add Serper credits or correct the API key, then rerun.", flush=True)
        return 2
    except Exception as exc:
        print(f"Search-provider preflight encountered a transient error; normal retry logic will continue: {type(exc).__name__}: {exc}", flush=True)

    with ThreadPoolExecutor(max_workers=worker_count) as pool:
        futures = {
            pool.submit(resolve_one, lic, name, broker, broker_license, api_key, cfg, cache, cache_path, audit_path): (lic, name, broker, broker_license)
            for lic, name, broker, broker_license in pending
        }
        try:
            for future in as_completed(futures):
                lic, name, broker, broker_license = futures[future]
                try:
                    result = future.result()
                except SearchProviderFatalError as exc:
                    set_provider_fatal(str(exc))
                    print(f"::error::Fatal search-provider failure: {exc}", flush=True)
                    interrupted = True
                    for pending_future in futures:
                        pending_future.cancel()
                    break
                except Exception as exc:
                    result = Result(
                        lic,
                        name,
                        broker,
                        status="Error",
                        notes=f"{type(exc).__name__}: {exc}",
                        updated_at=now_utc(),
                    )
                previous = checkpoint.get(result.license_number)
                attempted = finalize_result(result, previous)
                selected = choose_result(previous, attempted, force_replace=False)
                checkpoint[result.license_number] = selected
                append_history(history_path, previous, attempted, selected, args.mode)
                completed += 1

                now_tick = time.monotonic()
                if completed == 1 or completed == planned or completed % 10 == 0 or now_tick - last_progress_print >= 60:
                    elapsed = max(0.001, now_tick - run_started)
                    rate = completed / elapsed
                    remaining = max(0, planned - completed)
                    eta_seconds = int(remaining / rate) if rate > 0 else 0
                    print(
                        f"Progress: {completed}/{planned} ({completed / planned:.1%}); "
                        f"rate={rate:.2f} records/sec; remaining={remaining}; ETA={eta_seconds // 3600:02d}:{(eta_seconds % 3600) // 60:02d}:{eta_seconds % 60:02d}",
                        flush=True,
                    )
                    last_progress_print = now_tick

                if completed % checkpoint_every == 0 or completed == planned:
                    persist_all(
                        checkpoint_path, metadata_path, cache_path, input_file, output_file,
                        checkpoint, cache, completed, planned, args.mode,
                    )

                if interrupted:
                    for pending_future in futures:
                        pending_future.cancel()
                    break
        finally:
            # This final save covers ordinary exceptions and manual cancellation signals.
            persist_all(
                checkpoint_path, metadata_path, cache_path, input_file, output_file,
                checkpoint, cache, completed, planned, args.mode,
            )

    counts: dict[str, int] = defaultdict(int)
    for result in checkpoint.values():
        counts[result.status] += 1
    print("Done:", dict(counts), flush=True)
    if _provider_fatal.is_set():
        return 2
    return 130 if interrupted else 0

if __name__ == "__main__":
    raise SystemExit(main())
